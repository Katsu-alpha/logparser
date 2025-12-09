#
#   dif-client-table.py
#
#   Create a csv report of time series deltas for each client statistics
#
#   show ap debug client-table の出力を解析し、各クライアントの時系列のリトライ率、PHY rate, SNRをCSV形式で出力
#

import sys
import re
from collections import defaultdict
from aos_parser import AOSParser
import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

xlsfile = "client-rr.xlsx"

def column_str(n):
    """ Convert number to Excel column string (0 -> A, 26 -> AA) """
    s = ""
    while n >= 0:
        s = chr(n % 26 + ord('A')) + s
        n = n // 26 - 1
    return s

if len(sys.argv) < 2:
    print("Usage: dif-client-table.py <filename>...")
    exit(1)

fn = sys.argv[1]
f = open(fn, 'r', encoding='utf-8')
try:
    lines = f.readlines()
except UnicodeDecodeError:
    f = open(fn, 'r', encoding='macroman')
    lines = f.readlines()

p_txpkts = defaultdict(int)
p_txretr = defaultdict(int)
retry_rate = defaultdict(dict)
txrate = defaultdict(dict)
snr = defaultdict(dict)
times = set()
tim = None
in_cont = False
cmd = "show ap debug client-table"
buf = None
macs = set()
p_tim = ""

dic_txpkts = defaultdict(dict)
dic_txretr = defaultdict(dict)
dic_txpkts_d = defaultdict(dict)
dic_txretr_d = defaultdict(dict)

for l in lines:

    if in_cont:
        if l.startswith('Num '):
            in_cont = False
            if not tim:
                continue

            buf.append(l)
            # print("".join(buf))
            aos = AOSParser("".join(buf), [cmd])
            tbl = aos.get_table(cmd)

            if tbl is None:
                continue

            times.add(tim)

            for r in tbl[1:]:
                mac = r[0]
                if not re.match(r"[a-f0-9:]{17}$", mac):
                    continue
                macs.add(mac)
                txpkts = int(r[9])
                txretr = int(r[12])
                _txrate = int(r[13])
                _snr = int(r[16])

                txpkts_d = txpkts - p_txpkts[mac]
                txretr_d = txretr - p_txretr[mac]
                if txpkts_d < 0 or txretr_d < 0:
                    # counter wrap
                    txpkts_d = txpkts
                    txretr_d = txretr
                p_txpkts[mac] = txpkts
                p_txretr[mac] = txretr
                rr = f"{(txretr_d / txpkts_d * 100):.1f}" if txpkts_d > 1000 else ""
                retry_rate[mac][tim] = rr
                txrate[mac][tim] = _txrate
                snr[mac][tim] = _snr
                #
                dic_txpkts[mac][tim] = txpkts
                dic_txretr[mac][tim] = txretr
                dic_txpkts_d[mac][tim] = txpkts_d
                dic_txretr_d[mac][tim] = txretr_d

        else:
            buf.append(l)
            continue


    if l.startswith('TIME='):
        tim = l[5:].strip()
        continue

    if l.startswith('Client Table'):
        if tim[:5] == p_tim[:5]:
            continue
        # new minute
        p_tim = tim
        in_cont = True
        buf = [cmd+"\n", l]
        continue



#
#   Generate summary in CSV format
#
# print("Time,", end='')
# for m in sorted(macs):
    # print(f'"{m}",PHY rate,SNR,', end='')
    # print(f'"{m}",', end='')
# print()

tbl = []
for t in sorted(times):
    row = [t]
    # print(f"{t},", end="")
    for m in sorted(macs):
        rr = retry_rate[m].get(t, "")
        if rr != "NA" and rr != "":
            # rr = rr + "%" + f" {dic_txretr[m][t]}(+{dic_txretr_d[m][t]}) / {dic_txpkts[m][t]}(+{dic_txpkts_d[m][t]})"
            rr = rr + "%" + f" ({dic_txretr_d[m][t]} / {dic_txpkts_d[m][t]})"
        tr = txrate[m].get(t, "")
        s = snr[m].get(t, "")
        # print(f'"{rr}",{tr},{s},', end="")
        # print(f'"{rr}",', end="")
        row.append(rr)
    tbl.append(row)
    # print()


#
#   Create Excel
#
hdrs = ['Time']
hdrs.extend(sorted(macs))
df = pd.DataFrame(tbl, columns=hdrs)

#
#   Create Excel
#
wb = Workbook()
ws = wb.active
for r in dataframe_to_rows(df, index=False, header=True):
    ws.append(r)

f = Font(name='Aptos Narrow', size=9)
for row in ws.iter_rows(min_row=1):
    for cell in row:
        cell.font = f

widths = [15] * len(hdrs)
widths[0] = 8

for i,w in enumerate(widths):
    ws.column_dimensions[column_str(i)].width = w

f = Font(name='Aptos Narrow', bold=True, size=9)
Ses = PatternFill(fgColor="BDD7EE", fill_type="solid")
for cell in ws['A1':f'{column_str(len(hdrs)-1)}1'][0]:
    cell.fill = Ses
    cell.font = f


# go through all cells
thresh = [10.0, 20.0, 30.0, 40.0]
colors = ["63BE7B", "CCE3AB", "FFEB84", "FBAA77", "F8696B"]  # green, yellow, orange, red
for row in ws.iter_rows(min_row=2, min_col=2):
    for cell in row:
        v = cell.value
        if '%' in v:
            rr = float(v.split('%')[0])
            if rr > 100.0:
                cell.value = ""
                continue
            for i, t in enumerate(thresh):
                if rr < t:
                    cell.fill = PatternFill(fgColor=colors[i], fill_type="solid")
                    break
            else:
                cell.fill = PatternFill(fgColor=colors[-1], fill_type="solid")

# ws.auto_filter.ref = "A:G"
ws.freeze_panes = "B2"


#
#   output to file
#
print(f"Writing to {xlsfile} ... ", end="")
wb.save(xlsfile)
print("done.")

sys.exit(0)
